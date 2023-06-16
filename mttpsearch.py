from googlesearch import search
from datetime import date
from bs4 import BeautifulSoup
from urllib.error import HTTPError
import time
import requests
import openpyxl

# Read term txt files into lists
def createList(fileLoc):
    termsFile = open(fileLoc, "r")
    termsData = termsFile.read()
    return termsData.split("\n")
clinTermsList = createList("terms/clinicalTerms.txt")
techTermsList = createList("terms/techTerms.txt")
jobTermsList = createList("terms/jobTerms.txt")
badTermComboList = createList("terms/badTermCombo.txt")
sitesList = createList("terms/sites.txt")
nonoTermsList = createList("terms/nonoTerms.txt")
alwaysTermsList = createList("terms/alwaysTerms.txt")


# Create XLSX
xlBook = openpyxl.Workbook()
linkedInSheet = xlBook.active
linkedInSheet.title = ("LinkedIn")
wellfoundSheet = xlBook.create_sheet("Wellfound")
othersSheet = xlBook.create_sheet("Others")
data = (("date", "company", "job title", "area", "link", "term used"),)
for entry in data:
        linkedInSheet.append(entry)
        wellfoundSheet.append(entry)
        othersSheet.append(entry)

# Initialize Variables
headers = {"User-agent":"Mozilla/5.0 (Windows NT 10.0) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36"}
today = date.today()
clinTerm = ""
techTerm = ""
jobTerm = ""
site = ""
query = f'"{clinTerm}" "{techTerm}" "{jobTerm}" "remote" -"research" -"trial" -"trials" -"pharmacy technician" after:2023-06-14 site:"linkedin.com"|site:"wellfound.com"'
noError = True

# Scrape LinkedIn Page
def scrapeLinkedIn(soup, i):
    data = ()
    title = soup.find("title")
    if title == None:
        None
    else:
        if "hiring" in title.string:
            company = title.string[0:title.string.find("hiring")]
            jobTitle = title.string[title.string.find("hiring")+7:title.string.find(" in ")]
            area = title.string[title.string.find(" in ")+4:title.string.find(" | LinkedIn")]
            link = i
            termUsed = f'{clinTerm} {techTerm} {jobTerm}'
            data += ((today, company, jobTitle, area, link, termUsed),)
    for entry in data:
        linkedInSheet.append(entry)

# Scrape Wellfound Page
def scrapeWellfound(soup, i):
    print(soup)
    wellfoundSheet.append(i)

#Append Other Sites
def appendOtherSites(i):
    othersSheet.append((today, "none", "none", "none", i, f'{clinTerm} {techTerm} {jobTerm}'))

# Get sites with query
def getSites(query):
    linksGrabbed = ("",)
    for i in search(query, tld="com", num=10, stop=10, pause=2):
        for link in linksGrabbed:
            if i == link:
                continue
            else:
                linksGrabbed += (i,)
                page = requests.get(i, headers=headers)
                soup = BeautifulSoup(page.content, "html.parser")
                if "www.linkedin.com" in i:
                    scrapeLinkedIn(soup, i)
                elif "wellfound.com" in i:
                    scrapeWellfound(soup, i)
                else:
                    appendOtherSites(i)
        

for termOne in clinTermsList:
    clinTerm = termOne
    for termTwo in techTermsList:
        techTerm = termTwo
        for termThree in jobTermsList:
            badTerm = False
            jobTerm = termThree
            query = f'"{clinTerm}" "{techTerm}" "{jobTerm}" "remote" -"research" -"trial" -"trials" -"pharmacy technician" after:{today} site:linkedin.com | site:wellfound.com'
            for term in badTermComboList:
                if f'{clinTerm} {techTerm} {jobTerm}' == term:
                    badTerm = True
                    break
            if  badTerm:
                continue
            else:
                try:
                    getSites(query)
                except HTTPError:
                       print("HTTPError occurred")
                       continue
                print(f'{termOne} {termTwo} {termThree} done')
                time.sleep(10)
            
xlBook.save('mttpJobSearch.xlsx')